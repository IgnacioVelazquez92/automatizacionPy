import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

# Cargar el archivo Excel
archivo_excel = pd.read_excel('input/NPS.xlsx', sheet_name='Base XDSL')

# Crear la tabla pivote
tabla_pivote = pd.pivot_table(
    data=archivo_excel,
    index="Lider",
    columns="(Grupo) 2_NPS_GROUP - En base a tu último contacto realizado, ¿qué probabilidad hay de que recomi...",
    values="2 - En base a tu último contacto realizado, ¿qué probabilidad hay de que recomi...",
    aggfunc='count'
).round(0)

# Imprimir la tabla pivote
print(tabla_pivote)

# Guardar la tabla pivote en un nuevo archivo Excel
tabla_pivote.to_excel('output/tabla_pivote.xlsx', sheet_name='NPS')

# Cargar el archivo Excel usando openpyxl
wb = load_workbook('output/tabla_pivote.xlsx')

# Asegurarse de que la pestaña exista
if 'NPS' in wb.sheetnames:
    pestania = wb['NPS']
    min_col = pestania.min_column
    max_col = pestania.max_column
    min_row = pestania.min_row
    max_row = pestania.max_row

    # Imprimir los rangos de las celdas
    print(min_col, max_col, min_row, max_row)
else:
    print("La pestaña 'NPS' no existe en el archivo.")

#grafico
barchart = BarChart()
data= Reference(pestania, min_col=min_col+1, max_col=max_col , min_row=min_row, max_row=max_row)
categorias = Reference(pestania,min_col=min_col, max_col=min_col , min_row=min_row+1, max_row=max_row)

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categorias)

pestania.add_chart(barchart,'B12')
barchart.title='NPS'
barchart.style= 5

wb.save('output/tabla_pivote.xlsx')